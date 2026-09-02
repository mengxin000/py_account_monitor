from __future__ import annotations

import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .report_data import ReportData


# The order sheet follows the C++ OrderRecord columns.  The profit sheet is
# intentionally smaller: it contains only metrics that exist in the Binance
# monitor and the per-symbol position data used for vibration P&L.
BLUE = "1F4E78"
LIGHT_BLUE = "D9EAF7"
WHITE = "FFFFFF"
BLACK = "000000"
GREEN = "E2F0D9"
RED = "FCE4D6"
UNMATCHED = "DDEBF7"
GRID = "A6A6A6"
THIN = Side(style="thin", color=GRID)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
HEADER_FILL = PatternFill("solid", fgColor=BLUE)
SECTION_FILL = PatternFill("solid", fgColor=LIGHT_BLUE)
MATCH_FILL = PatternFill("solid", fgColor=GREEN)
LOSS_FILL = PatternFill("solid", fgColor=RED)
UNMATCHED_FILL = PatternFill("solid", fgColor=UNMATCHED)

ORDER_HEADERS = [
    "交易对", "订单ID", "订单系统ID", "订单方向", "订单成交数量", "订单平均成交价格", "订单手续费",
    "对冲订单ID", "对冲订单系统ID", "对冲订单方向", "对冲订单成交数量", "对冲订单平均成交价格",
    "对冲订单手续费", "收益", "成交价差", "成交时间",
]
ORDER_WIDTHS = [24, 23, 16, 11, 15, 18, 14, 23, 23, 11, 17, 21, 15, 16, 14, 23]


def _float(value: Any, default: float = 0.0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _time_text(value: Any) -> str:
    if value in (None, ""):
        return ""
    try:
        ms = int(float(value))
        return datetime.fromtimestamp(ms / 1000, tz=timezone.utc).astimezone().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
    except (TypeError, ValueError, OSError, OverflowError):
        return str(value)


def _style(cell: Any, fill: PatternFill | None = None, *, bold: bool = False, white_font: bool = False) -> None:
    cell.alignment = CENTER
    cell.border = BORDER
    cell.fill = fill or PatternFill("solid", fgColor=WHITE)
    cell.font = Font(name="Microsoft YaHei", size=10, bold=bold, color=WHITE if white_font else BLACK)


def _clear(ws: Any) -> None:
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)


def _order_row(row: dict[str, Any]) -> list[Any]:
    quantity = _float(row.get("quantity"))
    order_symbol = str(row.get("current_symbol") or "").upper()
    hedge_symbol = str(row.get("match_symbol") or "").upper()
    pair = f"{order_symbol}_{hedge_symbol}" if order_symbol and hedge_symbol else order_symbol or hedge_symbol
    return [
        pair,
        row.get("current_id", ""), row.get("current_system_id", ""), row.get("current_side", ""), quantity,
        _float(row.get("current_price")), _float(row.get("current_fee")),
        row.get("match_id", ""), row.get("match_system_id", ""), row.get("match_side", ""), quantity,
        _float(row.get("match_price")), _float(row.get("match_fee")), _float(row.get("profit")),
        _float(row.get("offset")), _time_text(row.get("event_time_ms")),
    ]


def _unmatched_row(row: dict[str, Any]) -> list[Any]:
    return [
        str(row.get("symbol") or "").upper(),
        row.get("id", ""), row.get("system_id", ""), row.get("side", ""), _float(row.get("quantity")),
        _float(row.get("price")), _float(row.get("fee")), "", "", "", "", "", "", "", "",
        _time_text(row.get("time_ms")),
    ]


def _write_orders(ws: Any, data: ReportData) -> None:
    _clear(ws)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    ws.append(ORDER_HEADERS)
    for cell in ws[1]:
        _style(cell, HEADER_FILL, bold=True, white_font=True)
    for col, width in enumerate(ORDER_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 34

    rows: list[tuple[dict[str, Any], str]] = [
        *((record, "match") for record in data.matches),
        *((record, "unmatched") for record in data.unmatched),
    ]
    rows.sort(key=lambda item: (
        str(item[0].get("current_symbol") or item[0].get("match_symbol") or item[0].get("symbol") or "").upper(),
        _float(item[0].get("event_time_ms", item[0].get("time_ms", 0))),
    ))
    for record, kind in rows:
        ws.append(_order_row(record) if kind == "match" else _unmatched_row(record))
        fill = UNMATCHED_FILL if kind == "unmatched" else (LOSS_FILL if _float(record.get("profit")) < 0 else MATCH_FILL)
        for cell in ws[ws.max_row]:
            _style(cell, fill)
        for col in (*range(5, 8), *range(11, 16)):
            ws.cell(ws.max_row, col).number_format = "0.000000"


def _position_rows(summary: dict[str, Any]) -> list[dict[str, Any]]:
    return list(summary.get("vibration_breakdown") or [])


def _pct_parts(summary: dict[str, Any]) -> tuple[float, float, float]:
    values = [abs(_float(summary.get("trade_profit"))), abs(_float(summary.get("fee_profit"))), abs(_float(summary.get("volatility_profit")))]
    total = sum(values)
    return tuple(value / total * 100 if total else 0.0 for value in values)  # type: ignore[return-value]


def _write_profit_block(ws: Any, data: ReportData) -> None:
    summary = data.profit_summary
    matched = int(summary["matched_count"])
    unmatched = int(summary["unmatched_count"])
    total = matched + unmatched
    match_ratio = matched / total * 100 if total else 0.0
    start = 1 if ws.max_row == 1 and ws["A1"].value is None else ws.max_row + 2

    # Account state, useful for checking the actual-equity calculation.
    ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=8)
    title = ws.cell(start, 1, f"账户收益分布  {data.account_id}  {data.day}")
    _style(title, HEADER_FILL, bold=True, white_font=True)
    ws.row_dimensions[start].height = 26
    account_row = start + 1
    account_values = [
        ("账户", data.account_id), ("交易日", data.day),
        ("09:30基准权益", data.baseline_equity if data.baseline_equity is not None else 0.0),
        ("当前权益", data.current_equity if data.current_equity is not None else 0.0),
    ]
    for index, (label, value) in enumerate(account_values):
        col = 1 + index * 2
        ws.cell(account_row, col, label)
        ws.cell(account_row, col + 1, value)
        _style(ws.cell(account_row, col), SECTION_FILL, bold=True)
        _style(ws.cell(account_row, col + 1))
        if isinstance(value, (int, float)):
            ws.cell(account_row, col + 1).number_format = "0.0000000000"

    metric_header = start + 3
    metrics = [
        ("实际损益", summary["actual_profit"], "0.0000000000"),
        ("交易损益", summary["trade_profit"], "0.0000000000"),
        ("费率损益", summary["fee_profit"], "0.0000000000"),
        ("24h总成交量", summary["total_volume"], "0.0000000000"),
    ]
    stats = [
        ("总成交条数", int(summary["fill_count"]), "0"),
        ("配对条数", matched, "0"),
        ("未配对条数", unmatched, "0"),
        ("盈利条数", int(summary["profit_count"]), "0"),
        ("亏损条数", int(summary["loss_count"]), "0"),
    ]
    for col, (header, value, fmt) in enumerate(metrics, 1):
        ws.cell(metric_header, col, header)
        ws.cell(metric_header + 1, col, value)
        _style(ws.cell(metric_header, col), HEADER_FILL, bold=True, white_font=True)
        _style(ws.cell(metric_header + 1, col))
        ws.cell(metric_header + 1, col).number_format = fmt

    stats_header = metric_header + 3
    for col, (header, value, fmt) in enumerate(stats, 1):
        ws.cell(stats_header, col, header)
        ws.cell(stats_header + 1, col, value)
        _style(ws.cell(stats_header, col), HEADER_FILL, bold=True, white_font=True)
        _style(ws.cell(stats_header + 1, col))
        ws.cell(stats_header + 1, col).number_format = fmt

    # Keep the per-symbol trading summary visible in the workbook.  This is
    # independent of the removed HK/US and market-order statistics.
    summary_header = stats_header + 4
    symbol_titles = ["基础币", "成交事件", "配对次数", "配对数量", "配对收益", "Exposure匹配数量", "Exposure匹配收益", "Exposure剩余数量"]
    for col, title_text in enumerate(symbol_titles, 1):
        ws.cell(summary_header, col, title_text)
        _style(ws.cell(summary_header, col), HEADER_FILL, bold=True, white_font=True)
    symbol_rows = data.symbol_summary()
    if symbol_rows:
        for row_index, item in enumerate(symbol_rows, summary_header + 1):
            values = [item["symbol"], int(item["fill_count"]), int(item["match_count"]), item["matched_quantity"], item["profit"], item["exposure_quantity"], item["exposure_profit"], item["unmatched_quantity"]]
            for col, value in enumerate(values, 1):
                ws.cell(row_index, col, value)
                _style(ws.cell(row_index, col))
                if col >= 4:
                    ws.cell(row_index, col).number_format = "0.0000000000"
    else:
        ws.cell(summary_header + 1, 1, "暂无成交记录")
        ws.merge_cells(start_row=summary_header + 1, start_column=1, end_row=summary_header + 1, end_column=len(symbol_titles))
        _style(ws.cell(summary_header + 1, 1))

    for col, width in enumerate([22, 18, 18, 18, 18, 18, 18, 18], 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"


def _new_workbook() -> Any:
    wb = Workbook()
    orders = wb.active
    orders.title = "成交订单明细"
    wb.create_sheet("收益分布")
    return wb


def build_workbook(data: ReportData, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = None
    if path.exists():
        try:
            candidate = load_workbook(path)
            distribution = candidate["收益分布"] if "收益分布" in candidate.sheetnames else None
            # Migrate the previous four-sheet/legacy layout once.  New files
            # are recognized by the account title in A1.
            if candidate.sheetnames == ["成交订单明细", "收益分布"] and distribution is not None and str(distribution["A1"].value or "").startswith("账户收益分布"):
                wb = candidate
            else:
                wb = _new_workbook()
        except Exception:
            wb = _new_workbook()
    else:
        wb = _new_workbook()
    _write_orders(wb["成交订单明细"], data)
    _write_profit_block(wb["收益分布"], data)
    temp = path.with_suffix(path.suffix + ".tmp")
    wb.save(temp)
    try:
        os.replace(temp, path)
        return path
    except PermissionError:
        fallback = path.with_name(f"{path.stem}_cpp{path.suffix}")
        os.replace(temp, fallback)
        return fallback
